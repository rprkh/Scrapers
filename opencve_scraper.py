import csv
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
import openpyxl
import requests
import os
import threading
import sys
from openpyxl import load_workbook
from dataclasses import dataclass
import multiprocessing

@dataclass
class CONFIG:
    OPENCVE_DIRECTORY = "OpenCVE"

event2 = multiprocessing.Event()

def extract_col_md_2(driver, url, vendor):
    driver.get(url)
    try:
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CLASS_NAME, "cve-header")))
        soup = BeautifulSoup(driver.page_source, 'html.parser')
        cve_headers = soup.find_all('tr', class_='cve-header')
        for header in cve_headers:
            cve_id = header.find('td', class_='col-md-2').text.strip()
            updated_date = header.find('td', class_='col-md-2 text-center').text.strip()
            if cve_id and not is_cve_id_scraped(cve_id, vendor):
                scrape_cve_info(cve_id, updated_date, vendor)
    except TimeoutException:
        print(f"Timeout occurred while waiting for the table for {vendor}.")

def is_cve_id_scraped(cve_id, vendor):
    filename = f"{vendor}_cve_details.xlsx"
    if os.path.isfile(filename):
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            if row[0] == cve_id:
                return True
    return False

def scrape_cve_info(cve_id, updated_date, vendor):
    url = f"https://www.opencve.io/cve/{cve_id}"
    response = requests.get(url)
   
    if response.status_code == 200:
        soup = BeautifulSoup(response.content, 'html.parser')
       
        cve_data = {}
        for title_element in soup.find_all('h4', class_='panel-title'):
            title = title_element.text.strip()
            data = title_element.find('span').text.strip()
            cve_data[title] = data
        description_div = soup.find('div', class_='col-md-9').find('div', class_='box box-primary').find('div', class_='box-body')
        description = description_div.text.strip() if description_div else "Description not available"
        cve_data['Description'] = description
 
        # Writing to Excel file row by row
        filename = f"{CONFIG.OPENCVE_DIRECTORY}/{vendor}_cve_details.xlsx"
        if not os.path.isfile(filename):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["CVE", "Attack Vector", "Attack Complexity", "Priveleges Required", "User Interaction", "Confidentiality Impact", "Integrity Complexity", "Availability Impact", "Scope", "Description", "Updated_date"])
            wb.save(filename)
       
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
        ws.append([cve_id, *cve_data.values(), updated_date])
        wb.save(filename)  # Save after scraping each CVE

def get_number_of_pages_per_vendor(vendor_name):
    url = "https://www.opencve.io/cve?vendor=" + vendor_name
    response = requests.get(url)
    html_content = response.text

    soup = BeautifulSoup(html_content, 'html.parser')

    pagination_ul = soup.find('ul', class_='pagination')

    if pagination_ul:
        li_tags = pagination_ul.find_all('li')

        if len(li_tags) >= 2:
            second_last_li = li_tags[-2]
            no_of_pages = second_last_li.text.strip()
        else:
            pass
    else:
        no_of_pages = 1
    
    return no_of_pages

def scrape_vendor(vendor):
    no_of_pages_for_vendor = get_number_of_pages_per_vendor(vendor)

    chrome_options = Options()
    chrome_options.add_argument("--disable-popup-blocking")
    chrome_options.add_argument("--headless") 

    driver = webdriver.Chrome(options=chrome_options)
    
    base_url = f"https://www.opencve.io/cve?vendor={vendor}&page="

    for page_no in range(1, int(no_of_pages_for_vendor) + 1):
        page_url = base_url + str(page_no)
        extract_col_md_2(driver, page_url, vendor)
        print(f"{page_no} {vendor}")

    event2.set()
    driver.quit()

# Read vendor names from command-line arguments
# vendor_keywords = sys.argv[1:]

# # Create threads for each vendor
# threads = []
# for vendor in vendor_keywords:
#     thread = threading.Thread(target=scrape_vendor, args=(vendor,))
#     threads.append(thread)

# # Start all threads
# for thread in threads:
#     thread.start()

# # Wait for all threads to complete
# for thread in threads:
#     thread.join()

# print("Excel files for CVE details of each vendor have been created.")

def delete_redundant_columns(directory, xlsx_filepath):
    wb = load_workbook(directory + xlsx_filepath)
    sheet = wb.active

    columns_to_delete = ['L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S']
    for column in reversed(columns_to_delete):
        sheet.delete_cols(sheet[column+'1'].column, 1)

    wb.save(directory + xlsx_filepath)

# for filename in os.listdir(CONFIG.OPENCVE_DIRECTORY):
#     if filename.endswith(".xlsx"):
#         delete_redundant_columns(CONFIG.OPENCVE_DIRECTORY + "/", filename)
